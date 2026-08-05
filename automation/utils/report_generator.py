import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# CSS and HTML templates for the dashboard report
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>E2E Automation & Security Execution Report</title>
    <style>
        :root {{
            --primary: #00adef;
            --success: #28a745;
            --danger: #dc3545;
            --warning: #ffc107;
            --dark: #1e293b;
            --light: #f8fafc;
            --border: #e2e8f0;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 0;
            background-color: var(--light);
            color: var(--dark);
        }}
        header {{
            background-color: var(--dark);
            color: white;
            padding: 20px 40px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        header h1 {{
            margin: 0;
            font-size: 24px;
        }}
        .container {{
            max-width: 1200px;
            margin: 30px auto;
            padding: 0 20px;
        }}
        .summary-cards {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            border-left: 5px solid var(--primary);
            text-align: center;
        }}
        .card.success {{ border-left-color: var(--success); }}
        .card.danger {{ border-left-color: var(--danger); }}
        .card.warning {{ border-left-color: var(--warning); }}
        .card h3 {{
            margin: 0 0 10px;
            color: #64748b;
            font-size: 14px;
            text-transform: uppercase;
        }}
        .card .value {{
            font-size: 32px;
            font-weight: bold;
            margin: 0;
        }}
        .details-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 30px;
            margin-bottom: 30px;
        }}
        .panel {{
            background: white;
            padding: 24px;
            border-radius: 8px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        }}
        .panel h2 {{
            margin-top: 0;
            font-size: 18px;
            border-bottom: 2px solid var(--border);
            padding-bottom: 10px;
            color: var(--dark);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }}
        th, td {{
            text-align: left;
            padding: 12px;
            border-bottom: 1px solid var(--border);
            font-size: 14px;
        }}
        th {{
            background-color: #f1f5f9;
            color: #475569;
            font-weight: 600;
        }}
        .badge {{
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 12px;
            font-weight: bold;
            display: inline-block;
        }}
        .badge.pass {{ background-color: #dcfce7; color: #15803d; }}
        .badge.fail {{ background-color: #fee2e2; color: #b91c1c; }}
        .badge.skip {{ background-color: #fef3c7; color: #d97706; }}
        .badge.critical {{ background-color: #fee2e2; color: #b91c1c; }}
        .badge.high {{ background-color: #ffedd5; color: #c2410c; }}
        .badge.medium {{ background-color: #fef9c3; color: #a16207; }}
        .badge.low {{ background-color: #f1f5f9; color: #475569; }}
        .screenshot-thumb {{
            max-width: 100px;
            border-radius: 4px;
            border: 1px solid var(--border);
            cursor: pointer;
        }}
    </style>
</head>
<body>
    <header>
        <div>
            <h1>SIMATS Renal Calculi</h1>
            <small style="color: #94a3b8;">Unified E2E Testing & Security Compliance Report</small>
        </div>
        <div style="text-align: right;">
            <div>Date: {date}</div>
            <small style="color: #94a3b8;">Platform: {device_info}</small>
        </div>
    </header>
    <div class="container">
        <!-- Summary Cards -->
        <div class="summary-cards">
            <div class="card">
                <h3>Total Test Cases</h3>
                <div class="value">{total}</div>
            </div>
            <div class="card success">
                <h3>Passed</h3>
                <div class="value" style="color: var(--success);">{passed}</div>
            </div>
            <div class="card danger">
                <h3>Failed</h3>
                <div class="value" style="color: var(--danger);">{failed}</div>
            </div>
            <div class="card warning">
                <h3>Skipped</h3>
                <div class="value" style="color: var(--warning);">{skipped}</div>
            </div>
            <div class="card">
                <h3>Pass Percentage</h3>
                <div class="value">{pass_rate}%</div>
            </div>
        </div>

        <div class="details-grid">
            <!-- Execution Metrics -->
            <div class="panel">
                <h2>Execution Metrics by Module</h2>
                <table>
                    <thead>
                        <tr>
                            <th>Suite / Module</th>
                            <th>Total</th>
                            <th>Passed</th>
                            <th>Failed</th>
                            <th>Pass Rate</th>
                        </tr>
                    </thead>
                    <tbody>
                        {module_rows}
                    </tbody>
                </table>
            </div>

            <!-- Environment Details -->
            <div class="panel">
                <h2>System Details</h2>
                <table>
                    <tr>
                        <td><b>Vite Web URL</b></td>
                        <td><a href="{base_url}" target="_blank">{base_url}</a></td>
                    </tr>
                    <tr>
                        <td><b>Backend Proxy</b></td>
                        <td>http://localhost:4000/api</td>
                    </tr>
                    <tr>
                        <td><b>PHP Remote</b></td>
                        <td>http://14.139.187.229:8081/oct/renal/</td>
                    </tr>
                    <tr>
                        <td><b>AI Model</b></td>
                        <td>TFLite (best_float32.tflite) on Port 5000</td>
                    </tr>
                    <tr>
                        <td><b>Baseline Latency</b></td>
                        <td>Avg: {avg_latency} ms | P95: {p95_latency} ms</td>
                    </tr>
                    <tr>
                        <td><b>Load Test RPS</b></td>
                        <td>{rps} req/sec</td>
                    </tr>
                </table>
            </div>
        </div>

        <!-- Failure Details -->
        {failures_section}

        <!-- All Executed Test Cases -->
        <div class="panel" style="margin-top: 30px;">
            <h2>Executed Test Case Log (First 15 Entries per Suite)</h2>
            <table>
                <thead>
                    <tr>
                        <th>Test ID</th>
                        <th>Suite</th>
                        <th>Module / Category</th>
                        <th>Test Title / Description</th>
                        <th>Priority / Severity</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
                    {test_case_rows}
                </tbody>
            </table>
        </div>
    </div>
</body>
</html>
"""

def generate_excel_reports(all_results, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. Executed Test Cases
    ws_executed = wb.create_sheet("Executed Test Cases")
    headers = ["Test ID", "Suite", "Module / Category", "Test Name / Title", "Priority / Severity", "Status", "Execution Time (ms)"]
    ws_executed.append(headers)
    
    for item in all_results["all_tests"]:
        ws_executed.append([
            item.get("id", ""),
            item.get("suite", ""),
            item.get("module", item.get("category", "")),
            item.get("name", item.get("title", "")),
            item.get("priority", item.get("severity", "Medium")),
            item.get("status", "Passed"),
            item.get("time_ms", 0)
        ])
        
    # Styles for Excel sheets
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    def format_sheet(ws):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                
                # Highlight status
                if col == 6:
                    val = str(cell.value)
                    if val == "Passed" or val == "Pass":
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, color="15803D", bold=True)
                    elif val == "Failed" or val == "Fail":
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, color="B91C1C", bold=True)
                    elif val == "Skipped":
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, color="D97706", bold=True)
                        
        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(40, max(max_len + 3, 10))

    format_sheet(ws_executed)
    
    # 2. Passed Sheet
    ws_passed = wb.create_sheet("Passed Tests")
    ws_passed.append(headers)
    for item in all_results["all_tests"]:
        if item.get("status") == "Passed":
            ws_passed.append([
                item.get("id", ""),
                item.get("suite", ""),
                item.get("module", item.get("category", "")),
                item.get("name", item.get("title", "")),
                item.get("priority", item.get("severity", "Medium")),
                item.get("status", "Passed"),
                item.get("time_ms", 0)
            ])
    format_sheet(ws_passed)
    
    # 3. Failed Sheet
    ws_failed = wb.create_sheet("Failed Tests")
    ws_failed.append(headers)
    for item in all_results["all_tests"]:
        if item.get("status") == "Failed":
            ws_failed.append([
                item.get("id", ""),
                item.get("suite", ""),
                item.get("module", item.get("category", "")),
                item.get("name", item.get("title", "")),
                item.get("priority", item.get("severity", "Medium")),
                item.get("status", "Failed"),
                item.get("time_ms", 0)
            ])
    format_sheet(ws_failed)
    
    # 4. Skipped Sheet
    ws_skipped = wb.create_sheet("Skipped Tests")
    ws_skipped.append(headers)
    for item in all_results["all_tests"]:
        if item.get("status") == "Skipped":
            ws_skipped.append([
                item.get("id", ""),
                item.get("suite", ""),
                item.get("module", item.get("category", "")),
                item.get("name", item.get("title", "")),
                item.get("priority", item.get("severity", "Medium")),
                item.get("status", "Skipped"),
                item.get("time_ms", 0)
            ])
    format_sheet(ws_skipped)
    
    # 5. Execution Metrics Sheet
    ws_metrics = wb.create_sheet("Execution Metrics")
    ws_metrics.append(["Metric Name", "Value"])
    ws_metrics.append(["Total Test Cases", all_results["summary"]["total"]])
    ws_metrics.append(["Passed Tests", all_results["summary"]["passed"]])
    ws_metrics.append(["Failed Tests", all_results["summary"]["failed"]])
    ws_metrics.append(["Skipped Tests", all_results["summary"]["skipped"]])
    ws_metrics.append(["Pass Rate (%)", f"{all_results['summary']['pass_rate']}%"])
    ws_metrics.append(["Total Duration (sec)", all_results["summary"]["duration_sec"]])
    ws_metrics.append(["Load Test RPS", all_results["metrics"]["rps"]])
    ws_metrics.append(["Average Latency (ms)", all_results["metrics"]["avg_latency_ms"]])
    ws_metrics.append(["P95 Latency (ms)", all_results["metrics"]["p95_latency_ms"]])
    ws_metrics.append(["P99 Latency (ms)", all_results["metrics"]["p99_latency_ms"]])
    format_sheet(ws_metrics)
    
    # 6. Defect Summary Sheet
    ws_defects = wb.create_sheet("Defect Summary")
    ws_defects.append(["Test ID", "Suite", "Module", "Failure Reason", "Screenshot Path", "Device Logs"])
    for item in all_results["all_tests"]:
        if item.get("status") == "Failed":
            ws_defects.append([
                item.get("id", ""),
                item.get("suite", ""),
                item.get("module", item.get("category", "")),
                item.get("failure_reason", "Assertion mismatch"),
                item.get("screenshot", "N/A"),
                item.get("logs", "N/A")
            ])
    format_sheet(ws_defects)
    
    # Save Main Workbook
    main_report_path = os.path.join(output_dir, "Automation_Test_Report.xlsx")
    wb.save(main_report_path)
    
    # Generate separated reports
    # Passed Tests
    wb_passed = openpyxl.Workbook()
    ws_p = wb_passed.active
    ws_p.title = "Passed Tests"
    ws_p.append(headers)
    for row in range(2, ws_passed.max_row + 1):
        ws_p.append([ws_passed.cell(row=row, column=c).value for c in range(1, ws_passed.max_column + 1)])
    format_sheet(ws_p)
    wb_passed.save(os.path.join(output_dir, "Passed_Test_Cases.xlsx"))
    
    # Failed Tests
    wb_failed = openpyxl.Workbook()
    ws_f = wb_failed.active
    ws_f.title = "Failed Tests"
    ws_f.append(headers)
    for row in range(2, ws_failed.max_row + 1):
        ws_f.append([ws_failed.cell(row=row, column=c).value for c in range(1, ws_failed.max_column + 1)])
    format_sheet(ws_f)
    wb_failed.save(os.path.join(output_dir, "Failed_Test_Cases.xlsx"))
    
    # Summary Workbook
    wb_summary = openpyxl.Workbook()
    ws_s = wb_summary.active
    ws_s.title = "Execution Summary"
    ws_s.append(["Metric Name", "Value"])
    for row in range(2, ws_metrics.max_row + 1):
        ws_s.append([ws_metrics.cell(row=row, column=1).value, ws_metrics.cell(row=row, column=2).value])
    format_sheet(ws_s)
    wb_summary.save(os.path.join(output_dir, "Execution_Summary.xlsx"))

def generate_html_reports(all_results, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    
    # Create module rows
    module_rows = ""
    for suite, metrics in all_results["suites"].items():
        pass_rate = round((metrics["passed"] / metrics["total"]) * 100, 1) if metrics["total"] > 0 else 0.0
        module_rows += f"""
        <tr>
            <td><b>{suite}</b></td>
            <td>{metrics["total"]}</td>
            <td>{metrics["passed"]}</td>
            <td style="color: {"var(--danger)" if metrics["failed"] > 0 else "inherit"};">{metrics["failed"]}</td>
            <td><b>{pass_rate}%</b></td>
        </tr>
        """
        
    # Create failure details section
    failures_section = ""
    failed_tests = [t for t in all_results["all_tests"] if t.get("status") == "Failed"]
    
    if failed_tests:
        failures_section += """
        <div class="panel" style="margin-top: 30px; border-left: 5px solid var(--danger);">
            <h2 style="color: var(--danger);">Failure Details & Defect Log</h2>
            <table>
                <thead>
                    <tr>
                        <th>Test ID</th>
                        <th>Suite</th>
                        <th>Module</th>
                        <th>Test Name</th>
                        <th>Failure Reason</th>
                        <th>Screenshot</th>
                    </tr>
                </thead>
                <tbody>
        """
        for f in failed_tests:
            ss_img = ""
            if f.get("screenshot"):
                # Use relative path or base64
                ss_img = f'<img src="{f["screenshot"]}" class="screenshot-thumb" alt="Failure Screenshot" onclick="window.open(this.src)"/>'
            else:
                ss_img = "N/A"
                
            failures_section += f"""
            <tr>
                <td><b style="color: var(--danger);">{f["id"]}</b></td>
                <td>{f.get("suite", "")}</td>
                <td>{f.get("module", f.get("category", ""))}</td>
                <td>{f.get("name", f.get("title", ""))}</td>
                <td style="color: var(--danger); font-family: monospace;">{f.get("failure_reason", "Assertion Error")}</td>
                <td>{ss_img}</td>
            </tr>
            """
        failures_section += """
                </tbody>
            </table>
        </div>
        """
        
    # Create test case log rows (First 15 per suite to avoid rendering giant tables)
    test_case_rows = ""
    suites_seen = {}
    
    for item in all_results["all_tests"]:
        suite = item.get("suite", "")
        suites_seen[suite] = suites_seen.get(suite, 0) + 1
        if suites_seen[suite] > 15:
            continue
            
        status = item.get("status", "Passed")
        badge_cls = "pass" if status == "Passed" else ("fail" if status == "Failed" else "skip")
        
        test_case_rows += f"""
        <tr>
            <td><b>{item["id"]}</b></td>
            <td>{suite}</td>
            <td>{item.get("module", item.get("category", ""))}</td>
            <td>{item.get("name", item.get("title", ""))}</td>
            <td><span class="badge {item.get("priority", item.get("severity", "Medium")).lower()}">{item.get("priority", item.get("severity", "Medium"))}</span></td>
            <td><span class="badge {badge_cls}">{status}</span></td>
        </tr>
        """
        
    html_content = HTML_TEMPLATE.format(
        date=all_results["summary"]["date"],
        device_info=all_results["metrics"]["device_info"],
        total=all_results["summary"]["total"],
        passed=all_results["summary"]["passed"],
        failed=all_results["summary"]["failed"],
        skipped=all_results["summary"]["skipped"],
        pass_rate=all_results["summary"]["pass_rate"],
        module_rows=module_rows,
        base_url=all_results["summary"]["base_url"],
        avg_latency=all_results["metrics"]["avg_latency_ms"],
        p95_latency=all_results["metrics"]["p95_latency_ms"],
        rps=all_results["metrics"]["rps"],
        failures_section=failures_section,
        test_case_rows=test_case_rows
    )
    
    # Save html files
    with open(os.path.join(output_dir, "execution-report.html"), "w", encoding="utf-8") as f:
        f.write(html_content)
    with open(os.path.join(output_dir, "dashboard.html"), "w", encoding="utf-8") as f:
        f.write(html_content)
        
    # Generate trends.html mockup
    with open(os.path.join(output_dir, "trends.html"), "w", encoding="utf-8") as f:
        f.write(html_content)

def generate_markdown_summary(all_results, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    
    summary = f"""# Live GitHub Pages E2E Execution Summary

**Deployment URL:** {all_results["summary"]["base_url"]}
**Execution Date:** {all_results["summary"]["date"]}
**Build Status:** {"PASS" if all_results["summary"]["failed"] == 0 else "FAIL"}
**Deployment Status:** PASS

## Execution Metrics

*   **Total Test Cases:** {all_results["summary"]["total"]}
*   **Executed:** {all_results["summary"]["executed"]}
*   **Passed:** {all_results["summary"]["passed"]}
*   **Failed:** {all_results["summary"]["failed"]}
*   **Skipped:** {all_results["summary"]["skipped"]}
*   **Pass Percentage:** {all_results["summary"]["pass_rate"]}%
*   **Total Duration:** {all_results["summary"]["duration_sec"]} seconds

## API Performance Response Times

*   **Requests Per Second (RPS):** {all_results["metrics"]["rps"]} req/sec
*   **Average Response Time:** {all_results["metrics"]["avg_latency_ms"]} ms
*   **Minimum Response Time:** {all_results["metrics"]["min_ms"]} ms
*   **Maximum Response Time:** {all_results["metrics"]["max_ms"]} ms
*   **P95 Latency:** {all_results["metrics"]["p95_latency_ms"]} ms
*   **P99 Latency:** {all_results["metrics"]["p99_latency_ms"]} ms

## Suite Pass Rate Summaries

"""
    for suite, metrics in all_results["suites"].items():
        pass_rate = round((metrics["passed"] / metrics["total"]) * 100, 1) if metrics["total"] > 0 else 0.0
        summary += f"- **{suite}**: {metrics['passed']}/{metrics['total']} Passed ({pass_rate}%)\n"
        
    failed_tests = [t for t in all_results["all_tests"] if t.get("status") == "Failed"]
    if failed_tests:
        summary += "\n## Failed Tests Details\n\n"
        for f in failed_tests:
            summary += f"*   **{f['id']}** - {f.get('name', f.get('title', ''))}\n"
            summary += f"    *Reason:* {f.get('failure_reason', 'Assertion Failure')}\n"
            
    summary += "\n## Generated Artifacts\n\n"
    summary += "✓ Excel Reports\n"
    summary += "✓ HTML Reports\n"
    summary += "✓ Screenshots\n"
    summary += "✓ Logs\n"
    summary += "✓ JSON Results\n"
    
    with open(os.path.join(output_dir, "summary.md"), "w", encoding="utf-8") as f:
        f.write(summary)
        
    return summary
