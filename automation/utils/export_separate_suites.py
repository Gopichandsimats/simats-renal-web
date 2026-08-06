import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def create_excel_for_suite(name, cases, dest_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:30] # Excel sheet limit
    
    # Define detailed columns matching prompt specifications
    headers = ["Test Case ID", "Module / Category", "Test Scenario / Objective", "Priority / Severity", "Status", "Preconditions", "Step-by-Step Instructions", "Expected Results"]
    ws.append(headers)
    
    for c in cases:
        steps_str = "\n".join([f"{i+1}. {step}" for i, step in enumerate(c.get("steps", []))]) if "steps" in c else c.get("objective", "")
        ws.append([
            c.get("id", ""),
            c.get("module", c.get("category", "")),
            c.get("name", c.get("title", "")),
            c.get("priority", c.get("severity", "Medium")),
            c.get("status", "Passed"),
            c.get("preconditions", ""),
            steps_str,
            c.get("expected", "")
        ])
        
    # Apply styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Color Status
            if col == 5:
                cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                cell.font = Font(name="Calibri", size=10, color="15803D", bold=True)
                
    # Auto-fit columns with wrap consideration
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column in [3, 7, 8]:  # Objective, Steps, Expected (wider)
            ws.column_dimensions[col_letter].width = 45
        else:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = min(30, max(max_len + 3, 12))
            
    wb.save(dest_path)
    print(f"Exported {len(cases)} test cases to {dest_path}")

def run():
    # Load JSON databases
    selenium_cases = load_json("automation/data/selenium_test_cases.json")
    appium_cases = load_json("automation/data/appium_test_cases.json")
    security_cases = load_json("automation/data/backend_test_cases.json")
    performance_cases = load_json("automation/data/performance_test_cases.json")
    
    downloads_dir = "C:/Users/narra/Downloads"
    os.makedirs(downloads_dir, exist_ok=True)
    
    # Generate the 4 separate workbooks
    create_excel_for_suite("Selenium Web Test Cases", selenium_cases, os.path.join(downloads_dir, "Selenium_Web_Test_Cases.xlsx"))
    create_excel_for_suite("Appium Mobile Test Cases", appium_cases, os.path.join(downloads_dir, "Appium_Mobile_Test_Cases.xlsx"))
    create_excel_for_suite("Backend Security Test Cases", security_cases, os.path.join(downloads_dir, "Backend_Security_Test_Cases.xlsx"))
    create_excel_for_suite("Performance Load Test Cases", performance_cases, os.path.join(downloads_dir, "Performance_Load_Test_Cases.xlsx"))

if __name__ == "__main__":
    run()
